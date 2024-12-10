from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    BarChart,
    Reference,
    Series,)

print("Recuperando dados do 1o arquivo...")
dict_anos = {}


arquivo1 = load_workbook(filename="gastos.xlsx")
ws1 = arquivo1['gastos']
max_linhas = ws1.max_row

for i in range(2, max_linhas+1):
    dict_anos[ws1['A%s' % i].value]['gastos'] = ws1['B%s' % i].value, 'receita'

print("Recuperando dados do 2o arquivo...")

arquivo2 = load_workbook(filename="receita.xlsx")
ws2 = arquivo2['receita']
max_linhas = ws2.max_row

for i in range(2, max_linhas+1):
    dict_anos[ws2['A%s' % i].value]['receita'] = ws2['B%s' % i].value, 'receita'

print("Criando novo arquivo...")

wb = Workbook()
ws = wb.active

ws['A1'] = 'Ano'
ws['B1'] = 'Gastos'
ws['C1'] = 'Receita'

i = 2
for key, value in dict_anos.items():
    ws['A%s' % i].value = key
    ws['B%s' % i].value = value['gastos']
    ws['C%s' % i].value = value['receita']
    i += 1

print("Adicionando gráficos...")
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Gastos x Receita"
chart1.y_axis.title = "R$"
chart1.x_axis.title = "Ano"

anos = Reference(ws, min_col=2, min_col=3, min_row=1, max_row=i)
data = Reference(ws, min_col=1, min_row=2, max_row=i)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(anos)
chart1.shape = 4

ws.add_chart(chart1, "A%s" % (i+2))

print("Salvando arquivo...")
wb.save("resultados.xlsx")

print("Arquivo criado com sucesso!")